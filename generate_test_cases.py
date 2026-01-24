#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
単体テストケースExcel生成スクリプト
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxlがインストールされていません。以下のコマンドでインストールしてください:")
    print("pip install openpyxl")
    exit(1)

def create_test_cases_excel():
    """テストケースExcelファイルを作成"""
    wb = Workbook()
    
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # スタイル定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 1. ユーザー登録・ログイン機能
    ws1 = wb.create_sheet("1. ユーザー登録・ログイン")
    create_user_auth_sheet(ws1, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 2. 商品管理機能
    ws2 = wb.create_sheet("2. 商品管理")
    create_product_sheet(ws2, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 3. 注文処理機能
    ws3 = wb.create_sheet("3. 注文処理")
    create_order_sheet(ws3, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 4. スタイリスト申請機能
    ws4 = wb.create_sheet("4. スタイリスト申請")
    create_stylist_apply_sheet(ws4, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 5. お問い合わせ機能
    ws5 = wb.create_sheet("5. お問い合わせ")
    create_inquiry_sheet(ws5, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 6. カート・お気に入り機能
    ws6 = wb.create_sheet("6. カート・お気に入り")
    create_cart_favorite_sheet(ws6, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 7. メルマガ登録機能
    ws7 = wb.create_sheet("7. メルマガ登録")
    create_newsletter_sheet(ws7, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 8. 管理者機能
    ws8 = wb.create_sheet("8. 管理者機能")
    create_admin_sheet(ws8, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 9. スタイリスト機能
    ws9 = wb.create_sheet("9. スタイリスト機能")
    create_stylist_sheet(ws9, header_fill, header_font, border, center_alignment, left_alignment)
    
    # 10. お客様の声機能
    ws10 = wb.create_sheet("10. お客様の声")
    create_testimonial_sheet(ws10, header_fill, header_font, border, center_alignment, left_alignment)
    
    # ファイル保存
    filename = "単体テストケース一覧.xlsx"
    wb.save(filename)
    print(f"✅ テストケースExcelファイルを作成しました: {filename}")

def create_user_auth_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """ユーザー登録・ログイン機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    # ヘッダーのスタイル設定
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # ユーザー登録 - 単項目チェック
        ["1", "メールアドレス未入力", "単項目", "なし", "email: 空, password: 'test123', name: 'テスト太郎'", "エラー: メールアドレス、パスワード、名前は必須です", ""],
        ["2", "パスワード未入力", "単項目", "なし", "email: 'test@example.com', password: 空, name: 'テスト太郎'", "エラー: メールアドレス、パスワード、名前は必須です", ""],
        ["3", "名前未入力", "単項目", "なし", "email: 'test@example.com', password: 'test123', name: 空", "エラー: メールアドレス、パスワード、名前は必須です", ""],
        ["4", "パスワードが6文字未満", "単項目", "なし", "email: 'test@example.com', password: 'test1', name: 'テスト太郎'", "エラー: パスワードは6文字以上である必要があります", ""],
        ["5", "メールアドレス形式不正", "単項目", "なし", "email: 'invalid-email', password: 'test123', name: 'テスト太郎'", "登録失敗（メールアドレス形式チェック）", "フロントエンドでチェック"],
        ["6", "正常な登録", "単項目", "なし", "email: 'newuser@example.com', password: 'test123', name: 'テスト太郎'", "ユーザー登録が完了しました", ""],
        
        # ユーザー登録 - 相関チェック
        ["7", "既存メールアドレスで登録", "相関", "メールアドレス 'existing@example.com' が既に登録済み", "email: 'existing@example.com', password: 'test123', name: 'テスト太郎'", "エラー: このメールアドレスは既に登録されています", ""],
        ["8", "オプション項目ありで登録", "相関", "なし", "email: 'user@example.com', password: 'test123', name: 'テスト太郎', lastName: 'テスト', firstName: '太郎', phone: '090-1234-5678'", "ユーザー登録が完了し、オプション項目も保存される", ""],
        
        # ログイン - 単項目チェック
        ["9", "メールアドレス未入力", "単項目", "なし", "email: 空, password: 'test123'", "エラー: メールアドレスとパスワードを入力してください", ""],
        ["10", "パスワード未入力", "単項目", "なし", "email: 'test@example.com', password: 空", "エラー: メールアドレスとパスワードを入力してください", ""],
        ["11", "正常なログイン", "単項目", "ユーザーが登録済み", "email: 'test@example.com', password: 'test123'", "ログインに成功しました", ""],
        
        # ログイン - 相関チェック
        ["12", "存在しないメールアドレス", "相関", "なし", "email: 'nonexistent@example.com', password: 'test123'", "エラー: メールアドレスまたはパスワードが正しくありません", ""],
        ["13", "間違ったパスワード", "相関", "ユーザーが登録済み（パスワード: 'test123'）", "email: 'test@example.com', password: 'wrongpass'", "エラー: メールアドレスまたはパスワードが正しくありません", ""],
        ["14", "一時停止ユーザーのログイン", "相関", "ユーザーが一時停止状態（isSuspended: true）", "email: 'suspended@example.com', password: 'test123'", "ログイン拒否（要確認）", "実装状況により異なる"],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:  # No, テスト種別
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    # 列幅調整
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_product_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """商品管理機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # 商品取得
        ["1", "商品一覧取得", "単項目", "なし", "GET /api/products", "全商品の一覧が取得できる", ""],
        ["2", "商品詳細取得（有効なslug）", "単項目", "商品が存在", "GET /api/products/slug/valid-slug", "商品詳細が取得できる", ""],
        ["3", "商品詳細取得（存在しないslug）", "単項目", "なし", "GET /api/products/slug/invalid-slug", "404エラーまたは空の結果", ""],
        
        # 商品登録（管理者）
        ["4", "商品名未入力", "単項目", "管理者ログイン済み", "name: 空, price: '¥10,000', slug: 'test-product'", "エラー: 商品名は必須です", ""],
        ["5", "価格未入力", "単項目", "管理者ログイン済み", "name: 'テスト商品', price: 空, slug: 'test-product'", "エラー: 価格は必須です", ""],
        ["6", "slug未入力", "単項目", "管理者ログイン済み", "name: 'テスト商品', price: '¥10,000', slug: 空", "エラー: slugは必須です", ""],
        ["7", "正常な商品登録", "単項目", "管理者ログイン済み", "name: 'テスト商品', price: '¥10,000', slug: 'test-product', description: '説明', stock: 10", "商品が正常に登録される", ""],
        
        # 商品登録 - 相関チェック
        ["8", "重複するslugで登録", "相関", "slug 'existing-slug' が既に存在", "slug: 'existing-slug', name: '新商品'", "エラー: このslugは既に使用されています", ""],
        ["9", "在庫数が負の値", "相関", "管理者ログイン済み", "stock: -1", "エラーまたは0に補正", ""],
        
        # 商品更新
        ["10", "存在しない商品IDで更新", "相関", "管理者ログイン済み", "id: 'invalid-id', name: '更新商品'", "404エラー", ""],
        ["11", "正常な商品更新", "相関", "商品が存在、管理者ログイン済み", "id: 'valid-id', name: '更新された商品名'", "商品が正常に更新される", ""],
        
        # 商品削除
        ["12", "存在しない商品IDで削除", "相関", "管理者ログイン済み", "id: 'invalid-id'", "404エラー", ""],
        ["13", "正常な商品削除", "相関", "商品が存在、管理者ログイン済み", "id: 'valid-id'", "商品が正常に削除される", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_order_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """注文処理機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # 注文作成 - 単項目チェック
        ["1", "商品ID未指定", "単項目", "なし", "items: []", "エラー: 商品を選択してください", ""],
        ["2", "数量未指定", "単項目", "商品が存在", "items: [{id: 'product-id', quantity: 0}]", "エラー: 数量を指定してください", ""],
        ["3", "配送先氏名未入力", "単項目", "なし", "shipping: {lastName: 空, firstName: '太郎'}", "エラー: 配送先情報を入力してください", ""],
        ["4", "メールアドレス未入力", "単項目", "なし", "shipping: {email: 空}", "エラー: メールアドレスを入力してください", ""],
        ["5", "郵便番号未入力", "単項目", "なし", "shipping: {postalCode: 空}", "エラー: 郵便番号を入力してください", ""],
        ["6", "電話番号未入力", "単項目", "なし", "shipping: {phone: 空}", "エラー: 電話番号を入力してください", ""],
        
        # 注文作成 - 相関チェック
        ["7", "在庫不足の商品を注文", "相関", "商品の在庫が5、注文数量が10", "items: [{id: 'product-id', quantity: 10}], product.stock: 5", "エラー: 商品「XXX」の在庫が不足しています（残り5点）", ""],
        ["8", "存在しない商品IDで注文", "相関", "なし", "items: [{id: 'invalid-id', quantity: 1}]", "エラー: 商品「XXX」が見つかりません", ""],
        ["9", "正常な注文（ログインユーザー）", "相関", "ユーザーログイン済み、在庫あり", "items: [{id: 'valid-id', quantity: 2}], userId: 'user-id'", "注文が作成され、在庫が減算される", ""],
        ["10", "正常な注文（ゲスト）", "相関", "在庫あり", "items: [{id: 'valid-id', quantity: 1}], userId: null", "注文が作成される（userIdはnull）", ""],
        ["11", "注文後在庫が減算される", "相関", "商品在庫10、注文数量3", "items: [{id: 'product-id', quantity: 3}]", "注文後、商品在庫が7になる", ""],
        ["12", "送料計算（15,000円未満）", "相関", "合計金額が14,000円", "payment: {total: 14000}", "送料500円が加算される", ""],
        ["13", "送料計算（15,000円以上）", "相関", "合計金額が15,000円以上", "payment: {total: 15000}", "送料0円", ""],
        
        # 注文取得
        ["14", "ユーザーID指定で注文取得", "相関", "ユーザーがログイン済み", "GET /api/orders?userId=user-id", "該当ユーザーの注文一覧が取得できる", ""],
        ["15", "全注文取得（管理者）", "相関", "管理者ログイン済み", "GET /api/orders", "全注文の一覧が取得できる", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_stylist_apply_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """スタイリスト申請機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # スタイリスト申請 - 単項目チェック
        ["1", "名前未入力", "単項目", "なし", "name: 空, bio: '自己紹介', email: 'stylist@example.com'", "エラー: 名前、自己紹介、メールアドレスは必須です", ""],
        ["2", "自己紹介未入力", "単項目", "なし", "name: 'スタイリスト', bio: 空, email: 'stylist@example.com'", "エラー: 名前、自己紹介、メールアドレスは必須です", ""],
        ["3", "メールアドレス未入力", "単項目", "なし", "name: 'スタイリスト', bio: '自己紹介', email: 空", "エラー: 名前、自己紹介、メールアドレスは必須です", ""],
        ["4", "正常な申請", "単項目", "なし", "name: 'スタイリスト', bio: '自己紹介', email: 'new@example.com'", "スタイリスト登録申請を受け付けました", ""],
        
        # スタイリスト申請 - 相関チェック
        ["5", "既存スタイリストのメールアドレスで申請", "相関", "メールアドレス 'existing@example.com' が既にスタイリスト登録済み", "email: 'existing@example.com'", "エラー: このメールアドレスは既に登録されています", ""],
        ["6", "申請中のメールアドレスで再申請", "相関", "メールアドレス 'pending@example.com' で申請中（status: pending）", "email: 'pending@example.com'", "エラー: このメールアドレスで既に申請が送信されています。審査をお待ちください。", ""],
        ["7", "パスワードありで申請", "相関", "なし", "password: 'stylist123'", "パスワードがハッシュ化されて保存される", ""],
        ["8", "パスワードなしで申請", "相関", "なし", "password: null", "申請は受理される（パスワードは後で設定可能）", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_inquiry_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """お問い合わせ機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # お問い合わせ作成 - 単項目チェック
        ["1", "名前未入力", "単項目", "なし", "name: 空, email: 'test@example.com', inquiryType: 'styling', message: 'メッセージ'", "エラー: すべての必須項目を入力してください", ""],
        ["2", "メールアドレス未入力", "単項目", "なし", "name: 'テスト', email: 空, inquiryType: 'styling', message: 'メッセージ'", "エラー: すべての必須項目を入力してください", ""],
        ["3", "問い合わせ種別未選択", "単項目", "なし", "name: 'テスト', email: 'test@example.com', inquiryType: 空, message: 'メッセージ'", "エラー: すべての必須項目を入力してください", ""],
        ["4", "メッセージ未入力", "単項目", "なし", "name: 'テスト', email: 'test@example.com', inquiryType: 'styling', message: 空", "エラー: すべての必須項目を入力してください", ""],
        ["5", "メールアドレス形式不正", "単項目", "なし", "email: 'invalid-email'", "エラー: 有効なメールアドレスを入力してください", ""],
        ["6", "正常なお問い合わせ作成", "単項目", "なし", "name: 'テスト', email: 'test@example.com', inquiryType: 'styling', message: 'メッセージ'", "お問い合わせが作成される", ""],
        
        # お問い合わせ作成 - 相関チェック
        ["7", "ログインユーザーでお問い合わせ", "相関", "ユーザーがログイン済み", "userId: 'user-id', name: 'テスト', email: 'test@example.com'", "お問い合わせにuserIdが紐付けられる", ""],
        ["8", "スタイリスト指定でお問い合わせ", "相関", "スタイリストが存在", "stylistId: 'stylist-id', inquiryType: 'styling'", "お問い合わせにstylistIdが紐付けられる", ""],
        ["9", "ゲストでお問い合わせ", "相関", "ログインしていない", "userId: null, name: 'ゲスト', email: 'guest@example.com'", "お問い合わせが作成される（userIdはnull）", ""],
        
        # お問い合わせ取得
        ["10", "ユーザーID指定で取得", "相関", "ユーザーがログイン済み", "GET /api/inquiries?userId=user-id", "該当ユーザーのお問い合わせ一覧が取得できる", ""],
        ["11", "スタイリストID指定で取得", "相関", "スタイリストがログイン済み", "GET /api/inquiries?stylistId=stylist-id", "該当スタイリストのお問い合わせ一覧が取得できる", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_cart_favorite_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """カート・お気に入り機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # カート機能
        ["1", "商品をカートに追加", "単項目", "商品が存在", "productId: 'product-id', quantity: 1", "カートに商品が追加される", "フロントエンド実装"],
        ["2", "カートから商品を削除", "単項目", "カートに商品が存在", "productId: 'product-id'", "カートから商品が削除される", "フロントエンド実装"],
        ["3", "カート内商品の数量変更", "単項目", "カートに商品が存在", "productId: 'product-id', quantity: 3", "カート内の商品数量が更新される", "フロントエンド実装"],
        ["4", "在庫数を超える数量をカートに追加", "相関", "商品在庫が5", "quantity: 10", "エラーまたは最大在庫数までに制限", "フロントエンド実装"],
        ["5", "カートが空の状態でチェックアウト", "相関", "カートが空", "items: []", "エラー: カートが空です", ""],
        
        # お気に入り機能
        ["6", "商品をお気に入りに追加", "単項目", "ユーザーがログイン済み、商品が存在", "productId: 'product-id'", "お気に入りに追加される", "フロントエンド実装"],
        ["7", "お気に入りから商品を削除", "単項目", "お気に入りに商品が存在", "productId: 'product-id'", "お気に入りから削除される", "フロントエンド実装"],
        ["8", "既にお気に入りに追加済みの商品を再追加", "相関", "商品が既にお気に入りに存在", "productId: 'existing-product-id'", "エラーまたは無視される", "フロントエンド実装"],
        ["9", "ログインしていない状態でお気に入り追加", "相関", "ログインしていない", "productId: 'product-id'", "ログインを促すメッセージ", "フロントエンド実装"],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_newsletter_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """メルマガ登録機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # メルマガ登録 - 単項目チェック
        ["1", "メールアドレス未入力", "単項目", "なし", "email: 空", "エラー: メールアドレスを入力してください", ""],
        ["2", "メールアドレス形式不正", "単項目", "なし", "email: 'invalid-email'", "エラー: 有効なメールアドレスを入力してください", ""],
        ["3", "正常なメルマガ登録", "単項目", "なし", "email: 'new@example.com'", "メルマガ登録が完了する", ""],
        
        # メルマガ登録 - 相関チェック
        ["4", "既に登録済みのメールアドレス", "相関", "メールアドレス 'existing@example.com' が既に登録済み", "email: 'existing@example.com'", "エラーまたは既に登録済みメッセージ", ""],
        ["5", "ログインユーザーでメルマガ登録", "相関", "ユーザーがログイン済み", "email: 'user@example.com', userId: 'user-id'", "メルマガ登録にuserIdが紐付けられる", ""],
        ["6", "メルマガ配信停止", "相関", "メルマガ登録済み", "email: 'subscribed@example.com', isActive: false", "メルマガ配信が停止される", ""],
        ["7", "メルマガ配信再開", "相関", "メルマガ配信停止中", "email: 'unsubscribed@example.com', isActive: true", "メルマガ配信が再開される", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_admin_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """管理者機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # ユーザー管理
        ["1", "ユーザー一時停止", "単項目", "管理者ログイン済み、ユーザーが存在", "userId: 'user-id', reason: '規約違反'", "ユーザーが一時停止される（isSuspended: true）", ""],
        ["2", "ユーザー有効化", "単項目", "ユーザーが一時停止中", "userId: 'user-id'", "ユーザーが有効化される（isSuspended: false）", ""],
        ["3", "ユーザー削除", "単項目", "管理者ログイン済み、ユーザーが存在", "userId: 'user-id'", "ユーザーが削除される", ""],
        ["4", "存在しないユーザーIDで操作", "相関", "管理者ログイン済み", "userId: 'invalid-id'", "404エラー", ""],
        ["5", "管理者自身を削除", "相関", "管理者ログイン済み", "userId: 'admin-user-id'（自分自身）", "エラーまたは拒否", ""],
        
        # スタイリスト管理
        ["6", "スタイリスト申請承認", "単項目", "管理者ログイン済み、申請が存在", "applicationId: 'app-id', status: 'approved'", "申請が承認され、スタイリストが作成される", ""],
        ["7", "スタイリスト申請却下", "単項目", "管理者ログイン済み、申請が存在", "applicationId: 'app-id', status: 'rejected'", "申請が却下される（status: rejected）", ""],
        ["8", "スタイリストアカウント無効化", "単項目", "管理者ログイン済み、スタイリストが存在", "stylistId: 'stylist-id', isActive: false", "スタイリストアカウントが無効化される", ""],
        ["9", "スタイリストアカウント有効化", "単項目", "スタイリストが無効化中", "stylistId: 'stylist-id', isActive: true", "スタイリストアカウントが有効化される", ""],
        
        # 監査ログ
        ["10", "監査ログ記録（ユーザー削除）", "相関", "管理者がユーザーを削除", "action: 'delete', targetType: 'user'", "監査ログに記録される", ""],
        ["11", "監査ログ記録（ユーザー一時停止）", "相関", "管理者がユーザーを一時停止", "action: 'suspend', targetType: 'user'", "監査ログに記録される", ""],
        ["12", "監査ログ取得", "単項目", "管理者ログイン済み", "GET /api/admin/audit-logs", "監査ログ一覧が取得できる", ""],
        
        # 注文管理
        ["13", "注文ステータス更新", "単項目", "管理者ログイン済み、注文が存在", "orderNumber: 'ORD-xxx', status: 'shipped'", "注文ステータスが更新される", ""],
        ["14", "存在しない注文番号で更新", "相関", "管理者ログイン済み", "orderNumber: 'ORD-invalid'", "404エラー", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_stylist_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """スタイリスト機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # スタイリストログイン
        ["1", "スタイリストログイン（正常）", "単項目", "スタイリストが登録済み", "email: 'stylist@example.com', password: 'stylist123'", "ログインに成功する", ""],
        ["2", "スタイリストログイン（間違ったパスワード）", "相関", "スタイリストが登録済み", "email: 'stylist@example.com', password: 'wrong'", "エラー: メールアドレスまたはパスワードが正しくありません", ""],
        ["3", "無効化されたスタイリストでログイン", "相関", "スタイリストが無効化中（isActive: false）", "email: 'inactive@example.com', password: 'stylist123'", "ログイン拒否", ""],
        
        # スタイリストプロフィール編集
        ["4", "プロフィール更新（正常）", "単項目", "スタイリストログイン済み", "name: '更新された名前', bio: '更新された自己紹介'", "プロフィールが更新される", ""],
        ["5", "プロフィール更新（名前未入力）", "単項目", "スタイリストログイン済み", "name: 空", "エラー: 名前は必須です", ""],
        
        # お問い合わせ対応
        ["6", "お問い合わせに返信", "単項目", "スタイリストログイン済み、お問い合わせが存在", "inquiryId: 'inquiry-id', message: '返信メッセージ'", "返信が作成される", ""],
        ["7", "存在しないお問い合わせに返信", "相関", "スタイリストログイン済み", "inquiryId: 'invalid-id'", "404エラー", ""],
        ["8", "他のスタイリストのお問い合わせに返信", "相関", "スタイリストAがログイン、お問い合わせはスタイリストBに紐付け", "inquiryId: 'other-stylist-inquiry'", "エラーまたは拒否", ""],
        
        # スタイリスト評価
        ["9", "スタイリストに評価を投稿", "単項目", "ユーザーがログイン済み、スタイリストが存在", "stylistId: 'stylist-id', rating: 5, comment: '素晴らしい'", "評価が作成される", ""],
        ["10", "同じスタイリストに2回評価", "相関", "ユーザーが既にスタイリストを評価済み", "stylistId: 'stylist-id', rating: 4", "エラー: 既に評価済みです", ""],
        ["11", "評価が1-5の範囲外", "単項目", "ユーザーがログイン済み", "rating: 6", "エラー: 評価は1-5の範囲で入力してください", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

def create_testimonial_sheet(ws, header_fill, header_font, border, center_alignment, left_alignment):
    """お客様の声機能のテストケース"""
    headers = ["No", "テスト項目", "テスト種別", "前提条件", "入力データ", "期待結果", "備考"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    test_cases = [
        # お客様の声投稿
        ["1", "名前未入力", "単項目", "なし", "name: 空, comment: 'コメント'", "エラー: 名前を入力してください", ""],
        ["2", "コメント未入力", "単項目", "なし", "name: 'テスト', comment: 空", "エラー: コメントを入力してください", ""],
        ["3", "正常な投稿", "単項目", "なし", "name: 'テスト太郎', comment: '素晴らしい商品でした'", "お客様の声が投稿される（isApproved: false）", ""],
        
        # お客様の声投稿 - 相関チェック
        ["4", "ログインユーザーで投稿", "相関", "ユーザーがログイン済み", "userId: 'user-id', name: 'テスト', comment: 'コメント'", "お客様の声にuserIdが紐付けられる", ""],
        ["5", "ゲストで投稿", "相関", "ログインしていない", "userId: null, name: 'ゲスト', comment: 'コメント'", "お客様の声が投稿される（userIdはnull）", ""],
        
        # 管理者機能
        ["6", "お客様の声を承認", "単項目", "管理者ログイン済み、投稿が存在", "testimonialId: 'testimonial-id', isApproved: true", "お客様の声が承認される", ""],
        ["7", "お客様の声を削除", "単項目", "管理者ログイン済み、投稿が存在", "testimonialId: 'testimonial-id'", "お客様の声が削除される", ""],
        ["8", "承認済みのお客様の声を表示", "相関", "お客様の声が承認済み", "GET /api/testimonials", "承認済みのお客様の声のみ表示される", ""],
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20

if __name__ == "__main__":
    create_test_cases_excel()
